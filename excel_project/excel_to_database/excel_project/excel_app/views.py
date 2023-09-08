from django.shortcuts import render
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook
from .models import GenBankBranch
from django.conf import settings
import os


def upload_file(request):
    if request.method == 'POST' and request.FILES['file']:
        uploaded_file = request.FILES['file']
        fs = FileSystemStorage()
        filename = fs.save(uploaded_file.name, uploaded_file)

        # Load the XLSX file
        # workbook = load_workbook(filename=fs.url(filename))
        workbook = load_workbook(filename=os.path.join(settings.MEDIA_ROOT, filename))
        sheet = workbook.active

        # Iterate through rows and save data to the database
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name, status = row
            # Assuming 'status' is a string containing 'active' or 'inactive'
            # Check if the status is valid before creating the object
            # if status in ['active', 'inactive']:
            GenBankBranch.objects.create(name=name, status=status)

        return render(request, 'upload_success.html')

    return render(request, 'upload.html')